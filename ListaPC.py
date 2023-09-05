import openpyxl

book = openpyxl.Workbook()
print(book.sheetnames)

pc_page = book.create_sheet('Meus Computadores')

pc_page.append(['Eletrônico', 'Memória Ram', 'Preço'])
pc_page.append(['Computador 1', '8gb ram', 'R$ 2500.00'])
pc_page.append(['Computador 2', '16gb ram', 'R$ 5500.00'])
pc_page.append(['Computador 3', '32gb ram', 'R$ 8700.00'])
pc_page.append(['Computador 4', '4gb ram', 'R$ 1200.00'])
pc_page.append(['Computador 5', '2gb ram', 'R$ 800.00'])

book.save('Planilha de Computadores.xlsx')